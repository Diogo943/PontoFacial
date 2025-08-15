import subprocess


bibliotecas_requeridas = ['cmake','customtkinter','opencv-contrib-python','numpy','pillow','insightface','onnxruntime', 'openpyxl','opencv-python']

for biblioteca in bibliotecas_requeridas:
    try:
        __import__(biblioteca)
    except ImportError:
        subprocess.check_call(['python', '-m', 'pip', 'install', biblioteca])



import customtkinter as ctk
import cv2
import threading
import os
import numpy as np
import time
from PIL import Image
from insightface.app import FaceAnalysis
import datetime
import openpyxl






class PontoFacial():
    def __init__(self):
        self.root = ctk.CTk()
        
        # Configuração da interface
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        
        largura, altura = self.root.winfo_screenmmwidth(), self.root.winfo_screenheight()
        posx = int(largura/2) + 250
        self.root.geometry(f"600x700+{posx}+1")
        self.root.title("Ponto Facial")
        
        self.ultima_deteccao = {}
        self.caminho_imagem_rostos = {}
        self.capturando = False
        self.caminho_pasta = os.getcwd().replace('model','\\FOTOS_RECONHECIMENTO_FACIAL')
        self.caminho_registro_excel = os.getcwd().replace('model','\\REGISTROS\\registro.xlsx')
        self.nome = "Desconhecido"


        self.widgets()
        self.horario_data()
        self.alerta_usuario()

        # Inicializando InsightFace
        self.app = FaceAnalysis(name = 'buffalo_l')
        self.app.prepare(ctx_id=-1, det_size=(640, 640))
        
        
        self.rostos_conhecidos = {}
        self.carregar_rostos(self.caminho_pasta)
        ##########################################################################
        self.root.mainloop()
        ##########################################################################


    def path_fotos(self, nome):
            return f'C:\\PONTO_FACIAL\\SECUTY\\Foto_{nome}.jpg'

    def criar_registro(self):
        with open(self.caminho_registro , "a", newline="") as arquivo:
            escritor = csv.writer(arquivo)
            escritor.writerow(['Data', 'Hora', 'Nome', 'Tipo'])
        print(f"Registro criado")

    def criar_registro_excel(self,data, hora, nome, tipo):
        planilha_registro = openpyxl.load_workbook(self.caminho_registro_excel)
        if 'Registro' not in planilha_registro.sheetnames:
            planilha_registro.create_sheet('Registro')

        planilha_registro['Registro'].append([data, hora, nome, tipo])
        planilha_registro.save(self.caminho_registro_excel)
        print(f"Registro adicionado ao Excel: {data} {hora} {nome} {tipo}")
    

    
    def carregar_rostos(self, caminho_pasta):
        for nome_arquivo in os.listdir(caminho_pasta):
            if nome_arquivo.lower().endswith(('.png', '.jpg', '.jpeg')):
                caminho_imagem = os.path.join(caminho_pasta, nome_arquivo)
                img = cv2.imread(caminho_imagem)
                rostos = self.app.get(img)
                if rostos:
                    self.rostos_conhecidos[nome_arquivo.split('.')[0]] = rostos[0].normed_embedding
                    self.caminho_imagem_rostos[nome_arquivo.split('.')[0]] = caminho_imagem
    
    def exibir_video(self):

        self.capturando = True
        #Inicializando a captura de video
        cap = cv2.VideoCapture(0)
        
        if not cap.isOpened():
            print("Erro ao acessar a câmera")
            return
        
        while self.capturando:
            ret, frame = cap.read()
            if not ret:
                break
            
            rostos = self.app.get(frame)
            
            for rosto in rostos:
                self.nome = "Desconhecido"
                melhor_similaridade = 0.6  # Limite de similaridade
                
                for nome_salvo, emb_salva in self.rostos_conhecidos.items():
                    similaridade = np.dot(rosto.normed_embedding, emb_salva)
                    if similaridade > melhor_similaridade:
                        melhor_similaridade = similaridade
                        self.nome = nome_salvo
                
                x1, y1, x2, y2 = rosto.bbox.astype(int)
                cor = (0, 255, 0) if self.nome != "Desconhecido" else (0, 0, 255)
                cv2.rectangle(frame, (x1, y1), (x2, y2), cor, 2)
                cv2.putText(frame, self.nome, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2)
                
                if self.nome != "Desconhecido" and (self.nome not in self.ultima_deteccao or (time.time() - self.ultima_deteccao[self.nome]) > 10):
                    self.ultima_deteccao[self.nome] = time.time()
                    threading.Thread(target=self.aguardar_e_mostrar_confirmacao, args=(self.nome,)).start()
  
            cv2.imshow('Ponto Facial', frame)
      
        cap.release()
        cv2.destroyAllWindows()

    def iniciar_video(self):
        threading.Thread(target=self.exibir_video).start()
              
    def parar_video(self):
        self.capturando = False
    
    def horario_data(self):

        self.horario = ctk.CTkLabel(self.main_frame, text=f'{datetime.datetime.now().strftime("%H:%M")}',
                                     font=('Times New Roman', 70, 'bold'), anchor=ctk.CENTER)
        self.horario.place(relx = 0.35, rely = 0.03)

        self.horario.after(1000, self.horario_data)
        
        #dicionário de meses do ano
        meses = {
            1: 'janeiro', 2: 'fevereiro', 3: 'março'   ,  4: 'abril'  , 5: 'maio'    , 6: 'junho',
            7: 'julho'  , 8: 'agosto'   , 9: 'setembro', 10: 'outubro',11: 'novembro',12: 'dezembro'
            }

        #dicionário de dias da semana
        dias_semana = {
                    0: 'segunda-feira', 1: 'terça-feira', 2: 'quarta-feira',
                    3: 'quinta-feira' , 4: 'sexta-feira' , 5: 'sábado'       ,6: 'domingo'
                    }
        
        nome_dia = dias_semana[datetime.datetime.now().weekday()]
        dia = datetime.datetime.now().day
        nome_mes = meses[datetime.datetime.now().month]
        ano = datetime.datetime.now().year

        data = f'{nome_dia},{dia} {nome_mes[:3]}. {ano}'

        self.data = ctk.CTkLabel(self.main_frame, text= data, font=('Times New Roman', 25, 'bold'), anchor=ctk.CENTER)
        self.data.place(relx = 0.3, rely = 0.2)
    
    def widgets(self):
        self.main_frame = ctk.CTkFrame(self.root)
        self.main_frame.place(relx=0.01, rely=0.01, relwidth=0.98, relheight=0.98)

        self.label_titulo = ctk.CTkLabel(self.main_frame, text="\U0001F600",font=( 'times',250))
        self.label_titulo.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)
        
        self.botao_iniciar = ctk.CTkButton(self.main_frame, text="Iniciar", font=('Times New Roman', 30, 'bold'), command=self.iniciar_video)
        self.botao_iniciar.place(relx=0.5, rely=0.9, anchor=ctk.CENTER)

        self.label_criador = ctk.CTkLabel(self.main_frame,
        text = 'Desenvolvido por: Diogo Marinho - Graduando de Engenharia Civil', font=('Times New Roman', 12))
        self.label_criador.place(relx=0.01, rely=0.95)
        
    def aguardar_e_mostrar_confirmacao(self, nome):
        self.main_frame_4 = ctk.CTkFrame(self.root)
        self.main_frame_4.place(relx=0.01, rely=0.01, relwidth=0.98, relheight=0.98)


        self.label_aguarde = ctk.CTkLabel(
            self.main_frame_4, 
            text="Reconhecendo, por favor aguarde...", 
            font=('Times New Roman', 25, 'bold')
        )
        self.label_aguarde.place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

        self.progress_bar = ctk.CTkProgressBar(
            self.main_frame_4, 
            mode="determinate", 
            progress_color="blue",
            determinate_speed=0.52,
            width=400,
            height=20
        )
        self.progress_bar.place(relx=0.5, rely=0.5, anchor=ctk.CENTER)
        self.progress_bar.set(0)
        self.progress_bar.start()

        time.sleep(2)
        self.parar_video()
        time.sleep(1)
        self.mostrar_confirmacao(nome)
    
    def mostrar_confirmacao(self, nome):
        self.main_frame_4.place_forget()
        self.root.deiconify()
        self.root.focus_force()
        self.main_frame_2 = ctk.CTkFrame(self.root)
        self.main_frame_2.place(relx=0.01, rely=0.01, relwidth=0.98, relheight=0.98)

        if nome == 'Desconhecido':
            ctk.CTkLabel(
                self.main_frame_2, 
                text="Usuário Desconhecido", 
                font=('Times New Roman', 20, 'bold')
            ).place(relx=0.5, rely=0.5, anchor=ctk.CENTER)
            ctk.CTkButton(
                self.main_frame_2, 
                text="Voltar", 
                command=self.voltar_menu,
                font=('Times New Roman', 40, 'bold'),
                width= 100
            ).place(relx=0.5, rely=0.9, anchor=ctk.CENTER)

        else:
            # Carregar e manter referência da imagem
            self.confirmation_image = ctk.CTkImage(
                dark_image=Image.open(self.caminho_imagem_rostos[nome]), 
                light_image= Image.open(self.caminho_imagem_rostos[nome]),
                size=(100, 120)
            )
            
            ctk.CTkLabel(
                self.main_frame_2, 
                image=self.confirmation_image, 
                text="", 
                corner_radius=50
            ).place(relx=0.5, rely=0.1, anchor=ctk.CENTER)

            ctk.CTkLabel(
                self.main_frame_2, 
                text=f'Que bom te ver,\n{nome}!', 
                font=('Times New Roman', 20, 'bold')
            ).place(relx=0.5, rely=0.3, anchor=ctk.CENTER)

            self.var_confirmacao = ctk.StringVar(value='off')

            self.check_entrada = ctk.CTkCheckBox(
                self.main_frame_2, 
                text="Entrada", 
                variable=self.var_confirmacao,
                onvalue='Entrada',
                offvalue='off',
                width= 100,
                font= ('Times New Roman', 30)
            )
            self.check_entrada.place(relx=0.25, rely=0.6, anchor=ctk.CENTER)
            
            self.check_comeco_intervalo = ctk.CTkCheckBox(
                self.main_frame_2, 
                text="Começo do Intervalo", 
                variable=self.var_confirmacao,
                onvalue='Comeco Intervalo',
                offvalue='off',
                width= 100,
                font= ('Times New Roman', 30)
            )
            self.check_comeco_intervalo.place(relx=0.25, rely=0.7, anchor=ctk.CENTER)
            
            self.check_fim_intervalo = ctk.CTkCheckBox(
                self.main_frame_2, 
                text="Fim do Intervalo", 
                variable=self.var_confirmacao,
                onvalue='Fim Intervalo',
                offvalue='off',
                width= 100,
                font= ('Times New Roman', 30)
            )
            self.check_fim_intervalo.place(relx=0.75, rely=0.7, anchor=ctk.CENTER)

            self.check_saida = ctk.CTkCheckBox(
                self.main_frame_2, 
                text="Saída", 
                variable=self.var_confirmacao,
                onvalue='Saida',
                offvalue='off',
                width= 100,
                font= ('Times New Roman', 30)
            )
            self.check_saida.place(relx=0.75, rely=0.6, anchor=ctk.CENTER)

            ctk.CTkButton(
                self.main_frame_2, 
                text="Confirmar", 
                command=lambda: self.finalizar_confirmacao(nome),
                font=('Times New Roman', 40, 'bold'),
                width= 100
            ).place(relx=0.5, rely=0.9, anchor=ctk.CENTER)

    def finalizar_confirmacao(self, nome):
        self.parar_video()
        self.ticket_confirmacao(nome)

        self.main_frame_2.place_forget()


    def ponto_registrar(self, nome, tipo):
        data = datetime.datetime.now().strftime("%d-%m-%Y")
        hora = datetime.datetime.now().strftime("%H")
        minuto = datetime.datetime.now().strftime("%M")
        minuto = int(minuto) - 7 #retroceder 7min
        segundo = datetime.datetime.now().strftime("%S") 
        horario_ponto = f"{hora}:{minuto}:{segundo}"
        self.criar_registro_excel(data, horario_ponto, nome, tipo)

    def ticket_confirmacao(self, nome):
        
        self.main_frame_3 = ctk.CTkFrame(self.root)
        self.main_frame_3.place(relx=0.01, rely=0.01, relwidth=0.98, relheight=0.98)

        self.label_despedida = ctk.CTkLabel(
            self.main_frame_3, 
            text="", 
            font=('Times New Roman', 15, 'bold')
        )
        self.label_despedida.place(relx=0.5, rely=0.82, anchor=ctk.CENTER)

        if self.var_confirmacao.get() == 'Entrada':


            self.label_entrada = ctk.CTkLabel(
                self.main_frame_3, 
                text="ENTRADA", 
                font=('Times New Roman', 20, 'bold')
            )
            self.label_entrada.place(relx=0.5, rely=0.1, anchor=ctk.CENTER)

            self.label_nome = ctk.CTkLabel(
                self.main_frame_3, 
                text=nome, 
                font=('Times New Roman', 20, 'bold')
            )
            self.label_nome.place(relx=0.5, rely=0.2, anchor='center')

            self.label_data = ctk.CTkLabel(
                self.main_frame_3, 
                text=f'Data\n{datetime.datetime.now().strftime("%d/%m/%Y")}', 
                font=('Times New Roman', 20, 'bold'),
                anchor='w'
            )
            self.label_data.place(relx=0.2, rely=0.5, anchor= 'center')

            self.label_horario = ctk.CTkLabel(
                self.main_frame_3, 
                text=f'Hora\n{datetime.datetime.now().strftime("%H:%M:%S")}', 
                font=('Times New Roman', 20, 'bold'),
                anchor='w'
            )
            self.label_horario.place(relx=0.8, rely=0.5, anchor= 'center')

            self.label_despedida.configure(text="Nos vemos em breve! :) Tenha um bom dia!")


            self.botao_ok = ctk.CTkButton(
                self.main_frame_3, 
                text="OK", 
                command=lambda: self.main_frame_3.place_forget(),
                font=('Times New Roman', 40, 'bold'),
                width= 100
            )
            self.botao_ok.place(relx=0.5, rely=0.9, anchor=ctk.CENTER)
            self.ponto_registrar(nome, self.var_confirmacao.get())


        elif self.var_confirmacao.get() == 'Saida':
            self.label_entrada = ctk.CTkLabel(
                self.main_frame_3, 
                text="SAÍDA", 
                font=('Times New Roman', 20, 'bold')
            )
            self.label_entrada.place(relx=0.5, rely=0.1, anchor=ctk.CENTER)

            self.label_nome = ctk.CTkLabel(
                self.main_frame_3, 
                text=nome, 
                font=('Times New Roman', 20, 'bold')
            )
            self.label_nome.place(relx=0.5, rely=0.2, anchor='center')

            self.label_data = ctk.CTkLabel(
                self.main_frame_3, 
                text=f'Data\n{datetime.datetime.now().strftime("%d/%m/%Y")}', 
                font=('Times New Roman', 20, 'bold'),
                anchor='w'
            )
            self.label_data.place(relx=0.2, rely=0.5, anchor= 'center')

            self.label_horario = ctk.CTkLabel(
                self.main_frame_3, 
                text=f'Hora\n{datetime.datetime.now().strftime("%H:%M:%S")}', 
                font=('Times New Roman', 20, 'bold'),
                anchor='w'
            )
            self.label_horario.place(relx=0.8, rely=0.5, anchor= 'center')

            self.label_despedida.configure(text="Até amanhã! :) Bom descanso.")

            
            self.botao_ok = ctk.CTkButton(
                self.main_frame_3, 
                text="OK", 
                command=lambda: self.main_frame_3.place_forget(),
                font=('Times New Roman', 40, 'bold'),
                width= 100
            )
            self.botao_ok.place(relx=0.5, rely=0.9, anchor=ctk.CENTER)
            self.ponto_registrar(nome, self.var_confirmacao.get())
        
        elif self.var_confirmacao.get() == 'Comeco Intervalo':
            self.label_intervalo = ctk.CTkLabel(
                self.main_frame_3, 
                text="COMEÇO DO INTERVALO", 
                font=('Times New Roman', 20, 'bold')
            )
            self.label_intervalo.place(relx=0.5, rely=0.1, anchor=ctk.CENTER)

            self.label_nome = ctk.CTkLabel(
                self.main_frame_3, 
                text=nome, 
                font=('Times New Roman', 20, 'bold')
            )
            self.label_nome.place(relx=0.5, rely=0.2, anchor='center')

            self.label_data = ctk.CTkLabel(
                self.main_frame_3, 
                text=f'Data\n{datetime.datetime.now().strftime("%d/%m/%Y")}', 
                font=('Times New Roman', 20, 'bold'),
                anchor='w'
            )
            self.label_data.place(relx=0.2, rely=0.5, anchor= 'center')

            self.label_horario = ctk.CTkLabel(
                self.main_frame_3, 
                text=f'Hora\n{datetime.datetime.now().strftime("%H:%M:%S")}', 
                font=('Times New Roman', 20, 'bold'),
                anchor='w'
            )
            self.label_horario.place(relx=0.8, rely=0.5, anchor= 'center')

            self.label_despedida.configure(text="Até depois mais! :) Bom almoço.")

            
            self.botao_ok = ctk.CTkButton(
                self.main_frame_3, 
                text="OK", 
                command=lambda: self.main_frame_3.place_forget(),
                font=('Times New Roman', 40, 'bold'),
                width= 100
            )
            self.botao_ok.place(relx=0.5, rely=0.9, anchor=ctk.CENTER)
            self.ponto_registrar(nome, self.var_confirmacao.get())
        
        elif self.var_confirmacao.get() == 'Fim Intervalo':
            self.label_intervalo = ctk.CTkLabel(
                self.main_frame_3, 
                text="FIM DO INTERVALO", 
                font=('Times New Roman', 20, 'bold')
            )
            self.label_intervalo.place(relx=0.5, rely=0.1, anchor=ctk.CENTER)

            self.label_nome = ctk.CTkLabel(
                self.main_frame_3, 
                text=nome, 
                font=('Times New Roman', 20, 'bold')
            )
            self.label_nome.place(relx=0.5, rely=0.2, anchor='center')

            self.label_data = ctk.CTkLabel(
                self.main_frame_3, 
                text=f'Data\n{datetime.datetime.now().strftime("%d/%m/%Y")}', 
                font=('Times New Roman', 20, 'bold'),
                anchor='w'
            )
            self.label_data.place(relx=0.2, rely=0.5, anchor= 'center')

            self.label_horario = ctk.CTkLabel(
                self.main_frame_3, 
                text=f'Hora\n{datetime.datetime.now().strftime("%H:%M:%S")}', 
                font=('Times New Roman', 20, 'bold'),
                anchor='w'
            )
            self.label_horario.place(relx=0.8, rely=0.5, anchor= 'center')

            self.label_despedida.configure(text="Até depois mais! :) Bom almoço.")

            
            self.botao_ok = ctk.CTkButton(
                self.main_frame_3, 
                text="OK", 
                command=lambda: self.main_frame_3.place_forget(),
                font=('Times New Roman', 40, 'bold'),
                width= 100
            )
            self.botao_ok.place(relx=0.5, rely=0.9, anchor=ctk.CENTER)
            self.ponto_registrar(nome, self.var_confirmacao.get())

        elif self.var_confirmacao.get() == 'off':
            self.main_frame_3.place_forget()
            self.janela_aviso = ctk.CTkToplevel(self.root)
            self.janela_aviso.grab_set()
    
            self.janela_aviso.geometry('400x400+600+100')
            self.janela_aviso.title('AVISO!')
    
            self.label_aviso = ctk.CTkLabel(self.janela_aviso,text= 'Por favor, informe o tipo de registro.',
                                            font=('Times New Roman',18), anchor= 'w')
            self.label_aviso.place(relx = 0.5, rely = 0.5, anchor = ctk.CENTER)
            
            self.botao_o = ctk.CTkButton(self.janela_aviso, text = 'OK', font = ('Times New Roman',20, 'bold'),
                                        command= lambda: self.janela_aviso.destroy())
            self.botao_o.place(relx = 0.5, rely = 0.9, anchor = ctk.CENTER)
            

            #self.mostrar_confirmacao(nome)


    def alerta_usuario(self):
        self.janela_alerta = ctk.CTkToplevel(self.root)
        self.janela_alerta.grab_set()

        self.janela_alerta.geometry('500x500+600+100')
        self.janela_alerta.title('AVISO!')

        self.label_alerta = ctk.CTkLabel(self.janela_alerta,text= 'Por favor, retire os óculos, máscara ou qualquer outro objeto\nque possa dificultar o reconhecimento facial.',
                                          font=('Times New Roman',18), anchor= 'w')
        self.label_alerta.place(relx = 0.5, rely = 0.3, anchor = ctk.CENTER)
        
        self.label_alerta_ = ctk.CTkLabel(self.janela_alerta,text= 'Por favor, sente-se próximo ao sensor da câmera\ne olhe para ele.',
                                          font=('Times New Roman',22, 'bold'), anchor= 'w')
        self.label_alerta_.place(relx = 0.5, rely = 0.55, anchor = ctk.CENTER)
        

        self.botao_ = ctk.CTkButton(self.janela_alerta, text = 'OK', font = ('Times New Roman',20, 'bold'),
                                    command= lambda: self.janela_alerta.destroy())
        self.botao_.place(relx = 0.5, rely = 0.9, anchor = ctk.CENTER)
    
    def voltar_menu(self):
        self.janela_confirmacao.destroy()
        time.sleep(1)
        self.root.deiconify()

pointFace = PontoFacial()
