from openpyxl import load_workbook
import os

class Registro():
    def __init__(self):
        self.arquivo = load_workbook(f'{os.getcwd()}\\REGISTROS\\registro.xlsx')

    def getDados(self):
        for i in range(1,self.arquivo['Registro'].max_row):
            coluna_data = self.arquivo['Registro'].cell(row=i, column=1).value
            coluna_hora = self.arquivo['Registro'].cell(row=i, column=2).value
            coluna_nome = self.arquivo['Registro'].cell(row=i, column=3).value
            coluna_tipo = self.arquivo['Registro'].cell(row=i, column=4).value

            print(f'{i} - {coluna_data} - {coluna_hora} - {coluna_nome} - {coluna_tipo}')

    def getId(self, nome = '', data = '', hora = '' ):
        for i in range(1,self.arquivo['Registro'].max_row):
            coluna_data = self.arquivo['Registro'].cell(row=i, column=1).value
            coluna_hora = self.arquivo['Registro'].cell(row=i, column=2).value
            coluna_nome = self.arquivo['Registro'].cell(row=i, column=3).value
            coluna_tipo = self.arquivo['Registro'].cell(row=i, column=4).value

            if nome.upper() in coluna_nome.upper() and data.upper() in coluna_data.upper() and hora.upper() in coluna_hora.upper():
                print(f'{i} - {coluna_data} - {coluna_hora} - {coluna_nome} - {coluna_tipo}')

    def setDados(self, data = '', hora = '' , nome = '' , tipo = '' ):
        max_linha = self.arquivo['Registro'].max_row + 1
        self.arquivo['Registro'].cell(row = max_linha, column=1).value =  data
        self.arquivo['Registro'].cell(row = max_linha, column=2).value = hora
        self.arquivo['Registro'].cell(row = max_linha, column=3).value = nome
        self.arquivo['Registro'].cell(row = max_linha, column=4).value = tipo

        self.arquivo.save(f'{os.getcwd()}\\REGISTROS\\registro.xlsx')
        self.arquivo.close()

    def clearDados(self, linha):

        self.arquivo['Registro'].delete_rows(idx=linha)

        self.arquivo.save(f'{os.getcwd()}\\REGISTROS\\registro.xlsx')
        self.arquivo.close()


registro = Registro()


def executar():
    menu = '''  
    1 - Visualizar registros
    2 - Procurar registro por nome/data/hora
    3 - Informar um registro
    4 - Deletar um registro
    0 - Sair do programa'''


    sair = False
    while not sair:
        registro = Registro()

        print(menu)

        try:
            op = int(input('Informe uma opção:'))
            if op == 1:
                registro.getDados()

            elif op == 2:
                nome = input('Informe o nome do estagiário:')
                data = input('Informe o data do registro (opcional):')
                hora = input('Informe o hora do registro (opcional):')

                registro.getId(nome, data, hora)

            elif op == 3:
                nome = input('Informe o nome do estagiário (obrigatório):')
                data = input('Informe o data do registro (obrigatório):')
                hora = input('Informe o hora do registro (obrigatório):')
                tipo = input('Informe o tipo do registro (obrigatório):')

                registro.setDados(data, hora, nome, tipo)
                print('Registro informado com sucesso!')

            elif op == 4:
                linha = int(input('Informe o linha do registro:'))
                registro.clearDados(linha)
                print('Registro deletado com sucesso!')

            else:
                break

        except ValueError:
            raise 'Valor incorreto'

executar()